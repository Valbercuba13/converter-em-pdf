from PIL import Image
from docx import Document
from reportlab.pdfgen import canvas
from openpyxl import load_workbook
import os

def convert_image_to_pdf(input_path, output_path):
    img = Image.open(input_path)
    img.save(output_path, "PDF", resolution=100.0)

def convert_text_to_pdf(input_path, output_path):
    with open(input_path, "r") as text_file:
        content = text_file.read()

    pdf_canvas = canvas.Canvas(output_path)
    pdf_canvas.drawString(100, 800, content)
    pdf_canvas.save()

def convert_word_to_pdf(input_path, output_path):
    doc = Document(input_path)
    pdf_canvas = canvas.Canvas(output_path)

    for para in doc.paragraphs:
        pdf_canvas.drawString(100, 800, para.text)

    pdf_canvas.save()

def convert_excel_to_pdf(input_path, output_path):
    wb = load_workbook(input_path)
    pdf_canvas = canvas.Canvas(output_path)

    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            pdf_canvas.drawString(100, 800, " ".join(str(cell) for cell in row))

    pdf_canvas.save()

def convert_files_in_directory():
    # Caminhos absolutos para diretórios de entrada e saída ajuda do chatgpt
    input_directory = os.path.join(os.getcwd(), "entrada")
    output_directory = os.path.join(os.getcwd(), "saida")

    
    if not os.path.exists(input_directory):
        print(f"Diretório de entrada '{input_directory}' não encontrado.")
        return

    
    os.makedirs(output_directory, exist_ok=True)

    for filename in os.listdir(input_directory):
        if filename.lower().endswith(".jpeg"):
            convert_image_to_pdf(os.path.join(input_directory, filename),
                                 os.path.join(output_directory, f"{os.path.splitext(filename)[0]}.pdf"))
        elif filename.lower().endswith(".txt"):
            convert_text_to_pdf(os.path.join(input_directory, filename),
                                os.path.join(output_directory, f"{os.path.splitext(filename)[0]}.pdf"))
        elif filename.lower().endswith(".docx"):
            convert_word_to_pdf(os.path.join(input_directory, filename),
                                os.path.join(output_directory, f"{os.path.splitext(filename)[0]}.pdf"))
        elif filename.lower().endswith(".xlsx"):
            convert_excel_to_pdf(os.path.join(input_directory, filename),
                                 os.path.join(output_directory, f"{os.path.splitext(filename)[0]}.pdf"))

if __name__ == "__main__":
    convert_files_in_directory()